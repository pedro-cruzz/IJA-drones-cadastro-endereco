from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from app import db
from app.models import Usuario, Solicitacao
from sqlalchemy.exc import IntegrityError
from datetime import datetime, date
from flask import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
from datetime import datetime, date 
import tempfile
from sqlalchemy.orm import joinedload


print("--- ROTAS CARREGADAS COM SUCESSO ---")

bp = Blueprint('main', __name__)
@bp.app_template_filter('datetimeformat')
def datetimeformat(value, format='%d-%m-%y'):
    try:
        # tenta converter string do tipo "2025-12-09"
        return datetime.strptime(value, "%Y-%m-%d").strftime(format)
    except:
        return value  # se falhar, retorna como está

# --- Context Processor: Simula o 'current_user' para o HTML ---
@bp.context_processor
def inject_user():
    class MockUser:
        is_authenticated = 'user_id' in session
        name = session.get('user_nome')
        id = session.get('user_id')
        tipo_usuario = session.get('user_tipo')
    return dict(current_user=MockUser())

# --- DASHBOARD UVIS ---

@bp.route('/')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    # AJUSTE CHAVE: Se for admin, operario OU visualizar, redireciona para o painel de gestão
    if session.get('user_tipo') in ['admin', 'operario', 'visualizar']:
        return redirect(url_for('main.admin_dashboard'))

    try:
        user_id = int(session.get('user_id'))
    except (ValueError, TypeError):
        session.clear()
        flash('Sessão Inválida. Por favor, faça login novamente.', 'warning')
        return redirect(url_for('main.login'))

    # 1. Query Base: Pega os pedidos SÓ deste usuário
    query = Solicitacao.query.filter_by(usuario_id=user_id)

    # 2. Lógica do Filtro: Verifica se veio algo na URL (ex: ?status=PENDENTE)
    filtro_status = request.args.get('status')

    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    # 3. Lógica da Paginação:
    page = request.args.get("page", 1, type=int)

    paginacao = query.order_by(
        Solicitacao.data_criacao.desc()
    ).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        'dashboard.html',
        nome=session.get('user_nome'),
        solicitacoes=paginacao.items,
        paginacao=paginacao
    )

# --- PAINEL DE GESTÃO (Visualização para todos) ---
@bp.route('/admin')
def admin_dashboard():
    # AJUSTE CHAVE: Permite 'admin', 'operario' E 'visualizar'
    if 'user_id' not in session or session.get('user_tipo') not in ['admin', 'operario', 'visualizar']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('main.login'))
    
    # Flag para controlar a renderização dos botões de edição no template
    is_editable = session.get('user_tipo') in ['admin', 'operario']
    
    # --- Captura filtros enviados pelo GET ---
    filtro_status = request.args.get("status")
    filtro_unidade = request.args.get("unidade")
    filtro_regiao = request.args.get("regiao")

    # --- Query base: Necessário dar JOIN com Usuario para filtrar por nome/região ---
    query = Solicitacao.query.join(Usuario)
    
    # 🔑 APLICAÇÃO DOS FILTROS 🔑
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))
    # 🔑 FIM APLICAÇÃO DOS FILTROS 🔑

    page = request.args.get("page", 1, type=int)

    paginacao = query.order_by(
        Solicitacao.data_criacao.desc()
    ).paginate(page=page, per_page=6)

    # Injeta a data/hora atual (para evitar o erro 'now is undefined' se fosse usado)
    data_atual = datetime.now() 
    
    return render_template(
        'admin.html',
        pedidos=paginacao.items,
        paginacao=paginacao,
        is_editable=is_editable,
        now=data_atual
    )

@bp.route('/admin/exportar_excel')
def exportar_excel():
    # Permite APENAS admin e operario
    if 'user_id' not in session or session.get('user_tipo') not in ['admin', 'operario']:
        flash('Permissão negada para exportar.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # --- Captura filtros ---
    filtro_status = request.args.get("status")
    filtro_unidade = request.args.get("unidade")
    filtro_regiao = request.args.get("regiao")

    # Query base
    query = Solicitacao.query.join(Usuario)

    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    pedidos = query.order_by(Solicitacao.data_criacao.desc()).all()

    # --- CRIA EXCEL ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Solicitações"

    # Cabeçalho atualizado com ENDEREÇO ÚNICO
    headers = [
        "ID", "Unidade", "Região",
        "Data Agendada", "Hora",
        "Endereço Completo",       # <-- CAMPO ÚNICO
        "Latitude", "Longitude",
        "Foco", "Tipo Visita", "Altura",
        "Criadouro?", "Apoio CET?",
        "Observação",
        "Status", "Protocolo", "Justificativa"
    ]

    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalho
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Conteúdo
    row_num = 2
    for p in pedidos:

        # --- ENDEREÇO COMPLETO ---
        endereco_completo = (
            f"{p.logradouro or ''}, {getattr(p, 'numero', '')} - "
            f"{p.bairro or ''} - "
            f"{(p.cidade or '')}/{(p.uf or '')} - "
            f"{p.cep or ''}"
        )

        if getattr(p, 'complemento', None):
            endereco_completo += f" - {p.complemento}"

        # Booleans
        criadouro_txt = "SIM" if getattr(p, 'criadouro', None) else "NÃO"
        cet_txt = "SIM" if getattr(p, 'apoio_cet', None) else "NÃO"

        # Data formatada
        if p.data_agendamento:
            try:
                if isinstance(p.data_agendamento, (date, datetime)):
                    data_formatada = p.data_agendamento.strftime("%d-%m-%y")
                else:
                    data_formatada = datetime.strptime(str(p.data_agendamento), "%Y-%m-%d").strftime("%d-%m-%y")
            except ValueError:
                data_formatada = str(p.data_agendamento)
        else:
            data_formatada = ""

        # Linha completa
        row = [
            p.id,
            p.autor.nome_uvis,
            p.autor.regiao,
            data_formatada,
            p.hora_agendamento,

            endereco_completo,     # <-- CAMPO ÚNICO AQUI

            getattr(p, 'latitude', ''),
            getattr(p, 'longitude', ''),

            p.foco,
            getattr(p, 'tipo_visita', ''),
            getattr(p, 'altura_voo', ''),
            criadouro_txt,
            cet_txt,
            getattr(p, 'observacao', ''),
            p.status,
            p.protocolo,
            p.justificativa
        ]

        # Escreve na planilha
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        row_num += 1

    # Congela o cabeçalho
    ws.freeze_panes = "A2"

    # Ajuste automático de largura
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar arquivo
    return send_file(
        output,
        download_name="relatorio_solicitacoes.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --- ROTA DE ATUALIZAÇÃO SIMPLES (Admin/Operário) ---
@bp.route('/admin/atualizar/<int:id>', methods=['POST'])
def atualizar(id):
    # AJUSTE CHAVE: Permite APENAS 'admin' E 'operario'
    if session.get('user_tipo') not in ['admin', 'operario']:
        flash('Permissão negada para esta ação.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    pedido = Solicitacao.query.get_or_404(id)

    # Campos de Geo/Status:
    pedido.protocolo = request.form.get('protocolo')
    pedido.status = request.form.get('status')
    pedido.justificativa = request.form.get('justificativa')
    pedido.latitude = request.form.get('latitude')
    pedido.longitude = request.form.get('longitude')


    db.session.commit()
    flash('Pedido atualizado com sucesso!', 'success')

    return redirect(url_for('main.admin_dashboard'))

# --- ROTA DE EDIÇÃO COMPLETA (Admin) ---
@bp.route('/admin/editar_completo/<int:id>', methods=['GET', 'POST'], endpoint='admin_editar') 
def admin_editar_completo(id):
    # AJUSTE CHAVE: Permite APENAS 'admin'
    if session.get('user_tipo') != 'admin':
        flash('Permissão negada. Apenas administradores podem acessar esta página.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    pedido = Solicitacao.query.get_or_404(id)

    if request.method == 'POST':
        try:
            # 1. Datas e Foco
            data_str = request.form.get('data_agendamento')
            hora_str = request.form.get('hora_agendamento')

            pedido.data_agendamento = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            pedido.hora_agendamento = datetime.strptime(hora_str, '%H:%M').time() if hora_str else None
            pedido.foco = request.form.get('foco')

            # 2. Detalhes Operacionais
            pedido.tipo_visita = request.form.get('tipo_visita')
            pedido.altura_voo = request.form.get('altura_voo')
            
            # Booleans
            pedido.criadouro = request.form.get('criadouro') == 'sim'
            pedido.apoio_cet = request.form.get('apoio_cet') == 'sim'
            pedido.observacao = request.form.get('observacao')

            # 3. Localização
            pedido.cep = request.form.get('cep')
            pedido.logradouro = request.form.get('logradouro')
            pedido.numero = request.form.get('numero')
            pedido.bairro = request.form.get('bairro')
            pedido.cidade = request.form.get('cidade')
            pedido.uf = request.form.get('uf')
            pedido.complemento = request.form.get('complemento')
            
            # GPS
            pedido.latitude = request.form.get('latitude')
            pedido.longitude = request.form.get('longitude')

            # 4. Status e Decisão (Controle Interno)
            pedido.protocolo = request.form.get('protocolo')
            pedido.status = request.form.get('status')
            pedido.justificativa = request.form.get('justificativa')

            db.session.commit()
            flash('Solicitação atualizada (Edição Completa) com sucesso!', 'success')
            return redirect(url_for('main.admin_dashboard'))
        
        except ValueError as ve:
            db.session.rollback()
            flash(f"Erro no formato de data/hora: {ve}", "warning")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar: {e}", "danger")
    
    # Renderiza o formulário pré-preenchido
    return render_template('admin_editar_completo.html', pedido=pedido)

# --- NOVO PEDIDO ---
@bp.route('/novo_cadastro', methods=['GET', 'POST'], endpoint='novo')
def novo():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    hoje = date.today().isoformat()

    if request.method == 'POST':
        try:
            user_id_int = int(session['user_id'])

            data_str = request.form.get('data')
            hora_str = request.form.get('hora')

            if data_str:
                data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
            else:
                data_obj = None

            if hora_str:
                hora_obj = datetime.strptime(hora_str, '%H:%M').time()
            else:
                hora_obj = None

            criadouro_bool = request.form.get('criadouro') == 'sim'
            apoio_cet_bool = request.form.get('apoio_cet') == 'sim'


            nova_solicitacao = Solicitacao(
                data_agendamento=data_obj,
                hora_agendamento=hora_obj,

                cep=request.form.get('cep'),
                logradouro=request.form.get('logradouro'),
                bairro=request.form.get('bairro'),
                cidade=request.form.get('cidade'),
                numero=request.form.get('numero'),
                uf=request.form.get('uf'),
                complemento=request.form.get('complemento'), 

                foco=request.form.get('foco'),

                tipo_visita=request.form.get('tipo_visita'),
                altura_voo=request.form.get('altura_voo'),
                criadouro=criadouro_bool,
                apoio_cet=apoio_cet_bool,
                observacao=request.form.get('observacao'),

                latitude=request.form.get('latitude'),
                longitude=request.form.get('longitude'),

                usuario_id=user_id_int,
                status='PENDENTE'
            )

            db.session.add(nova_solicitacao)
            db.session.commit()

            flash('Pedido enviado!', 'success')
            return redirect(url_for('main.dashboard'))

        except ValueError as ve:
            db.session.rollback()
            flash(f"Erro no formato de data/hora: {ve}", "warning")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar: {e}", "danger")

    return render_template('cadastro.html', hoje=hoje)

# --- LOGIN ---
@bp.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        # AJUSTE CHAVE: Redireciona para admin_dashboard se for admin, operario OU visualizar
        if session.get('user_tipo') in ['admin', 'operario', 'visualizar']:
            return redirect(url_for('main.admin_dashboard'))
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        user = Usuario.query.filter_by(login=request.form.get('login')).first()

        if user and user.check_senha(request.form.get('senha')):
            session['user_id'] = int(user.id)
            session['user_nome'] = user.nome_uvis
            session['user_tipo'] = user.tipo_usuario

            flash(f'Bem-vindo, {user.nome_uvis}! Login realizado com sucesso.', 'success')

            # AJUSTE CHAVE: Redireciona para admin_dashboard se for admin, operario OU visualizar
            if user.tipo_usuario in ['admin', 'operario', 'visualizar']:
                return redirect(url_for('main.admin_dashboard'))
            return redirect(url_for('main.dashboard'))
        else:
            flash('Login ou senha incorretos. Tente novamente.', 'danger')

    return render_template('login.html')

# --- LOGOUT ---
@bp.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('main.login'))

@bp.route("/forcar_erro")
def forcar_erro():
    1 / 0  # erro proposital
    return "nunca vai chegar aqui"

# ReportLab (PDF)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)

# Openpyxl (Excel)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# O objeto 'bp' precisa ser definido (Exemplo: bp = Blueprint('main', __name__))
# E 'Usuario' e 'Solicitacao' precisam ser seus modelos SQLAlchemy

# =======================================================================
# Função Auxiliar de Filtros (Reutilizada em todas as rotas)
# =======================================================================

def aplicar_filtros_base(query, filtro_data, uvis_id):
    """Aplica o filtro de mês/ano e opcionalmente o filtro de UVIS (usuario_id)."""
    
    # Filtro de Mês/Ano (obrigatório)
    query = query.filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
    
    # Filtro de UVIS (opcional)
    if uvis_id:
        query = query.filter(Solicitacao.usuario_id == uvis_id)
        
    return query


# =======================================================================
# ROTA 1: Visualização do Relatório (HTML)
# =======================================================================
@bp.route('/relatorios', methods=['GET'])
def relatorios():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    # 1. Parâmetros de Filtro
    mes_atual = request.args.get('mes', datetime.now().month, type=int)
    ano_atual = request.args.get('ano', datetime.now().year, type=int)
    uvis_id = request.args.get('uvis_id', type=int)
    filtro_data = f"{ano_atual}-{mes_atual:02d}"

    # 2. UVIS disponíveis para o dropdown
    uvis_disponiveis = db.session.query(Usuario.id, Usuario.nome_uvis) \
        .filter(Usuario.tipo_usuario == 'uvis') \
        .order_by(Usuario.nome_uvis) \
        .all()

    # 3. Histórico Mensal (usado para gerar anos disponíveis - não filtra por uvis_id)
    dados_mensais_raw = (
        db.session.query(
            db.func.strftime('%Y-%m', Solicitacao.data_criacao).label('mes'),
            db.func.count(Solicitacao.id)
        )
        .group_by('mes')
        .order_by('mes')
        .all()
    )
    dados_mensais = [tuple(row) for row in dados_mensais_raw]

    anos_disponiveis = sorted(list(set([d[0].split('-')[0] for d in dados_mensais])), reverse=True)
    if not anos_disponiveis:
        anos_disponiveis = [ano_atual]

    # 4. Totalizações (usando a função de filtro)
    base_query = db.session.query(Solicitacao)

    total_solicitacoes = aplicar_filtros_base(base_query, filtro_data, uvis_id).count()
    
    total_aprovadas = aplicar_filtros_base(base_query, filtro_data, uvis_id) \
        .filter(Solicitacao.status == "APROVADO").count()

    total_recusadas = aplicar_filtros_base(base_query, filtro_data, uvis_id) \
        .filter(Solicitacao.status == "NEGADO").count()

    total_analise = aplicar_filtros_base(base_query, filtro_data, uvis_id) \
        .filter(Solicitacao.status == "EM ANÁLISE").count()

    total_pendentes = aplicar_filtros_base(base_query, filtro_data, uvis_id) \
        .filter(Solicitacao.status == "PENDENTE").count()

    # 5. Consultas Agrupadas (usando a função de filtro)

    # Região (requer JOIN)
    query_regiao = db.session.query(Usuario.regiao, db.func.count(Solicitacao.id)) \
        .join(Usuario, Usuario.id == Solicitacao.usuario_id)
    dados_regiao_raw = aplicar_filtros_base(query_regiao, filtro_data, uvis_id) \
        .group_by(Usuario.regiao) \
        .order_by(db.func.count(Solicitacao.id).desc()) \
        .all()
    dados_regiao = [tuple(row) for row in dados_regiao_raw]

    # Status
    query_status = db.session.query(Solicitacao.status, db.func.count(Solicitacao.id))
    dados_status_raw = aplicar_filtros_base(query_status, filtro_data, uvis_id) \
        .group_by(Solicitacao.status) \
        .order_by(db.func.count(Solicitacao.id).desc()) \
        .all()
    dados_status = [tuple(row) for row in dados_status_raw]

    # Foco
    query_foco = db.session.query(Solicitacao.foco, db.func.count(Solicitacao.id))
    dados_foco_raw = aplicar_filtros_base(query_foco, filtro_data, uvis_id) \
        .group_by(Solicitacao.foco) \
        .order_by(db.func.count(Solicitacao.id).desc()) \
        .all()
    dados_foco = [tuple(row) for row in dados_foco_raw]
    
    # Tipo Visita
    query_tipo_visita = db.session.query(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
    dados_tipo_visita_raw = aplicar_filtros_base(query_tipo_visita, filtro_data, uvis_id) \
        .group_by(Solicitacao.tipo_visita) \
        .order_by(db.func.count(Solicitacao.id).desc()) \
        .all()
    dados_tipo_visita = [tuple(row) for row in dados_tipo_visita_raw]
    
    # Altura de Voo
    query_altura_voo = db.session.query(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
    dados_altura_voo_raw = aplicar_filtros_base(query_altura_voo, filtro_data, uvis_id) \
        .group_by(Solicitacao.altura_voo) \
        .order_by(db.func.count(Solicitacao.id).desc()) \
        .all()
    dados_altura_voo = [tuple(row) for row in dados_altura_voo_raw]

    # Unidade (UVIS) - Requer JOIN e filtro adicional de tipo_usuario
    query_unidade = db.session.query(Usuario.nome_uvis, db.func.count(Solicitacao.id)) \
        .join(Usuario, Usuario.id == Solicitacao.usuario_id) \
        .filter(Usuario.tipo_usuario == 'uvis')
    dados_unidade_raw = aplicar_filtros_base(query_unidade, filtro_data, uvis_id) \
        .group_by(Usuario.nome_uvis) \
        .order_by(db.func.count(Solicitacao.id).desc()) \
        .all()
    dados_unidade = [tuple(row) for row in dados_unidade_raw]

    # 6. Retorno
    return render_template(
        'relatorios.html',
        total_solicitacoes=total_solicitacoes,
        total_aprovadas=total_aprovadas,
        total_recusadas=total_recusadas,
        total_analise=total_analise,
        total_pendentes=total_pendentes,
        dados_regiao=dados_regiao,
        dados_status=dados_status,
        dados_foco=dados_foco,
        dados_tipo_visita=dados_tipo_visita,
        dados_altura_voo=dados_altura_voo,
        dados_unidade=dados_unidade,
        dados_mensais=dados_mensais,
        mes_selecionado=mes_atual,
        ano_selecionado=ano_atual,
        anos_disponiveis=anos_disponiveis,
        uvis_id_selecionado=uvis_id, # Passa o ID selecionado
        uvis_disponiveis=uvis_disponiveis # Passa a lista completa para o dropdown
    )


# =======================================================================
# ROTA 2: Exportar PDF (Com Filtro UVIS)
# =======================================================================
import os
import tempfile
from io import BytesIO
from datetime import datetime
from math import ceil

from flask import send_file, request
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage, Flowable, KeepTogether
)

# matplotlib é opcional — tentamos importar e marcamos se disponível
try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except Exception:
    MATPLOTLIB_AVAILABLE = False

@bp.route('/admin/exportar_relatorio_pdf')
def exportar_relatorio_pdf():
    # -------------------------
    # 1. Parâmetros e filtros
    # -------------------------
    mes = int(request.args.get('mes', datetime.now().month))
    ano = int(request.args.get('ano', datetime.now().year))
    uvis_id = request.args.get('uvis_id', type=int)
    orient = request.args.get('orient', default='portrait')  # 'portrait' ou 'landscape'
    filtro_data = f"{ano}-{mes:02d}"

    # 2. Busca Principal para Totais e Detalhes (mesma lógica sua)
    query_base = db.session.query(Solicitacao, Usuario).join(Usuario, Usuario.id == Solicitacao.usuario_id)
    query_base = aplicar_filtros_base(query_base, filtro_data, uvis_id)
    query_results = query_base.order_by(Solicitacao.data_criacao.desc()).all()

    # 3. Totais
    total_solicitacoes = len(query_results)
    total_aprovadas = sum(1 for s, u in query_results if s.status == "APROVADO")
    total_recusadas = sum(1 for s, u in query_results if s.status == "NEGADO")
    total_analise = sum(1 for s, u in query_results if s.status == "EM ANÁLISE")
    total_pendentes = sum(1 for s, u in query_results if s.status == "PENDENTE")

    # 4. Buscas agrupadas (mantendo sua lógica)
    def aplicar_filtros_agrupados(query):
        query = query.filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
        if uvis_id:
            query = query.filter(Solicitacao.usuario_id == uvis_id)
        return query

    dados_regiao_raw = aplicar_filtros_agrupados(
        db.session.query(Usuario.regiao, db.func.count(Solicitacao.id)).join(Usuario, Usuario.id == Solicitacao.usuario_id)
    ).group_by(Usuario.regiao).all()
    dados_regiao = [(r or "Não informado", c) for r, c in dados_regiao_raw]

    dados_status_raw = aplicar_filtros_agrupados(
        db.session.query(Solicitacao.status, db.func.count(Solicitacao.id))
    ).group_by(Solicitacao.status).all()
    dados_status = [(s or "Não informado", c) for s, c in dados_status_raw]

    dados_foco_raw = aplicar_filtros_agrupados(
        db.session.query(Solicitacao.foco, db.func.count(Solicitacao.id))
    ).group_by(Solicitacao.foco).all()
    dados_foco = [(f or "Não informado", c) for f, c in dados_foco_raw]

    dados_tipo_visita_raw = aplicar_filtros_agrupados(
        db.session.query(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
    ).group_by(Solicitacao.tipo_visita).all()
    dados_tipo_visita = [(t or "Não informado", c) for t, c in dados_tipo_visita_raw]

    dados_altura_raw = aplicar_filtros_agrupados(
        db.session.query(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
    ).group_by(Solicitacao.altura_voo).all()
    dados_altura_voo = [(a or "Não informado", c) for a, c in dados_altura_raw]

    dados_unidade_query = db.session.query(Usuario.nome_uvis, db.func.count(Solicitacao.id)) \
        .join(Usuario, Usuario.id == Solicitacao.usuario_id) \
        .filter(Usuario.tipo_usuario == 'uvis')
    dados_unidade_raw = aplicar_filtros_agrupados(dados_unidade_query).group_by(Usuario.nome_uvis).order_by(db.func.count(Solicitacao.id).desc()).all()
    dados_unidade = [(u or "Não informado", c) for u, c in dados_unidade_raw]

    dados_mensais_raw = (
        db.session.query(
            db.func.strftime('%Y-%m', Solicitacao.data_criacao).label('mes'),
            db.func.count(Solicitacao.id)
        )
        .group_by('mes')
        .order_by('mes')
        .all()
    )
    dados_mensais = [(m, c) for m, c in dados_mensais_raw]

    # -------------------------
    # 5. Preparar documento PDF
    # -------------------------
    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    caminho_pdf = tmp_pdf.name
    tmp_pdf.close()

    pagesize = A4
    if orient == 'landscape':
        pagesize = landscape(A4)

    doc = SimpleDocTemplate(caminho_pdf,
                            pagesize=pagesize,
                            leftMargin=16*mm, rightMargin=16*mm,
                            topMargin=16*mm, bottomMargin=20*mm)

    # Styles aprimorados
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=22, leading=26, alignment=1, spaceAfter=8, textColor=colors.HexColor('#0d6efd'))
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666'), alignment=1, spaceAfter=6)
    section_h = ParagraphStyle('sec', parent=styles['Heading2'], fontSize=12, spaceAfter=6, textColor=colors.HexColor('#0d6efd'))
    normal = styles['Normal']
    small = ParagraphStyle('small', parent=styles['BodyText'], fontSize=9, textColor=colors.HexColor('#555'))

    story = []

    # -------------------------
    # Funções utilitárias
    # -------------------------
    def safe_img_from_plt(fig):
        """Recebe um matplotlib.figure.Figure, retorna ReportLab Image (BytesIO)."""
        bio = BytesIO()
        fig.tight_layout()
        fig.savefig(bio, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig)
        bio.seek(0)
        return RLImage(bio, width=170*mm)  # escala automática

    def render_small_table(rows, colWidths):
        tbl = Table(rows, colWidths=colWidths)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fbfdff')]),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        return tbl

    # -------------------------
    # Cabeçalho / Capa
    # -------------------------
    # Logo: procura em static/logo.png por padrão — se não existir, pula
    logo_path = os.path.join(os.getcwd(), 'static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=36*mm, height=36*mm)
        except Exception:
            logo = None
    else:
        logo = None

    # Título e capa
    story.append(Spacer(1, 6))
    if logo:
        # coloca o logo e título lado a lado
        h = [[logo, Paragraph(f"<b>Relatório Mensal — {mes:02d}/{ano}</b>", title_style)]]
        cap_tbl = Table(h, colWidths=[40*mm, (doc.width - 40*mm)])
        cap_tbl.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
        story.append(cap_tbl)
    else:
        story.append(Paragraph(f"Relatório Mensal — {mes:02d}/{ano}", title_style))

    # subtítulo e linhas
    titulo_uvis = ""
    if uvis_id:
        uvis_obj = db.session.query(Usuario.nome_uvis).filter(Usuario.id == uvis_id).first()
        if uvis_obj:
            titulo_uvis = f" — {uvis_obj.nome_uvis}"
    story.append(Paragraph(f"Sistema de Gestão de Solicitações{titulo_uvis}", subtitle_style))
    story.append(Spacer(1, 8))

    # capa: box com resumo principal (centralizado)
    resumo_box = [
        ['Métrica', 'Quantidade'],
        ['Total de Solicitações', str(total_solicitacoes)],
        ['Aprovadas', str(total_aprovadas)],
        ['Recusadas', str(total_recusadas)],
        ['Em Análise', str(total_analise)],
        ['Pendentes', str(total_pendentes)]
    ]
    story.append(render_small_table(resumo_box, [80*mm, 40*mm]))
    story.append(Spacer(1, 12))

    # Capa: breve meta-infos
    story.append(Paragraph(f"Gerado por: Sistema SGSV — Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", small))
    story.append(Spacer(1, 18))

    # -------------------------
    # Sumário simples (lista de seções)
    # -------------------------
    story.append(Paragraph("Sumário", section_h))
    sumario_itens = [
        "Resumo Geral",
        "Solicitações por Região",
        "Status Detalhado",
        "Solicitações por Foco / Tipo / Altura",
        "Solicitações por Unidade (UVIS)",
        "Histórico Mensal",
        "Gráficos (Visão Geral)",
        "Registros Detalhados"
    ]
    for i, it in enumerate(sumario_itens, 1):
        story.append(Paragraph(f"{i}. {it}", normal))
    story.append(PageBreak())

    # -------------------------
    # Seções com tabelas (formatadas)
    # -------------------------
    # 1) Resumo Geral (repetição do box com estilo)
    story.append(Paragraph("Resumo Geral", section_h))
    story.append(render_small_table(resumo_box, [110*mm, 50*mm]))
    story.append(Spacer(1, 8))

    # 2) Regiões
    story.append(Paragraph("Solicitações por Região", section_h))
    rows = [['Região', 'Total']] + [[r, str(c)] for r, c in dados_regiao]
    story.append(render_small_table(rows, [110*mm, 50*mm]))
    story.append(Spacer(1, 8))

    # 3) Status
    story.append(Paragraph("Status Detalhado", section_h))
    rows = [['Status', 'Total']] + [[s, str(c)] for s, c in dados_status]
    story.append(render_small_table(rows, [110*mm, 50*mm]))
    story.append(Spacer(1, 8))

    # 4) Foco / Tipo / Altura
    story.append(Paragraph("Solicitações por Foco", section_h))
    rows = [['Foco', 'Total']] + [[f, str(c)] for f, c in dados_foco]
    story.append(render_small_table(rows, [110*mm, 50*mm]))
    story.append(Spacer(1, 6))

    story.append(Paragraph("Solicitações por Tipo de Visita", section_h))
    rows = [['Tipo', 'Total']] + [[t, str(c)] for t, c in dados_tipo_visita]
    story.append(render_small_table(rows, [110*mm, 50*mm]))
    story.append(Spacer(1, 6))

    story.append(Paragraph("Solicitações por Altura de Voo", section_h))
    rows = [['Altura (m)', 'Total']] + [[str(a), str(c)] for a, c in dados_altura_voo]
    story.append(render_small_table(rows, [110*mm, 50*mm]))
    story.append(Spacer(1, 8))

    # 5) UVIS
    story.append(Paragraph("Solicitações por Unidade (UVIS) — Top", section_h))
    rows = [['Unidade', 'Total']] + [[u, str(c)] for u, c in dados_unidade]
    story.append(render_small_table(rows, [110*mm, 50*mm]))
    story.append(Spacer(1, 8))

    # 6) Histórico mensal
    story.append(Paragraph("Histórico Mensal (Total por Mês)", section_h))
    rows = [['Mês', 'Total']] + [[m, str(c)] for m, c in dados_mensais]
    story.append(render_small_table(rows, [70*mm, 40*mm]))
    story.append(Spacer(1, 12))

    # 7) Gráficos — somente se matplotlib disponível
    story.append(PageBreak())
    story.append(Paragraph("Gráficos (Visão Geral)", section_h))
    if MATPLOTLIB_AVAILABLE:
        try:
            # Pie chart: distribuição por status
            labels = [s for s, _ in dados_status]
            values = [c for _, c in dados_status]
            fig1, ax1 = plt.subplots(figsize=(6, 3))
            ax1.pie(values or [1], labels=labels, autopct=lambda p: f'{p:.0f}%' if p > 0 else '', startangle=90, textprops={'fontsize': 8})
            ax1.axis('equal')
            story.append(safe_img_from_plt(fig1))
            story.append(Spacer(1, 8))

            # Bar chart: top UVIS
            u_names = [u for u, _ in dados_unidade[:8]]
            u_vals = [c for _, c in dados_unidade[:8]]
            fig2, ax2 = plt.subplots(figsize=(8, 2.6))
            ax2.barh(u_names[::-1] or ['Nenhum'], u_vals[::-1] or [0])
            ax2.set_xlabel('Total')
            ax2.set_title('Top UVIS (maiores)', fontsize=9)
            ax2.tick_params(axis='y', labelsize=8)
            story.append(safe_img_from_plt(fig2))
            story.append(Spacer(1, 8))

            # Line chart: histórico mensal
            months = [m for m, _ in dados_mensais]
            counts = [c for _, c in dados_mensais]
            fig3, ax3 = plt.subplots(figsize=(8, 2.6))
            if months:
                ax3.plot(months, counts, marker='o', linewidth=1)
                ax3.set_xticklabels(months, rotation=45, fontsize=8)
            ax3.set_title('Histórico Mensal', fontsize=9)
            ax3.grid(axis='y', linestyle=':', linewidth=0.5)
            story.append(safe_img_from_plt(fig3))
            story.append(Spacer(1, 6))
        except Exception:
            # se algo falhar nos gráficos, apenas passa
            story.append(Paragraph("Gráficos indisponíveis (erro ao gerar).", normal))
            story.append(Spacer(1, 8))
    else:
        story.append(Paragraph("Matplotlib não disponível — gráficos foram omitidos.", normal))
        story.append(Spacer(1, 8))

    # 8) Registros detalhados (tabela grande)
    story.append(PageBreak())
    story.append(Paragraph("Registros Detalhados", section_h))
    story.append(Spacer(1, 6))

    registros_header = ['Data', 'Hora', 'Unidade', 'Protocolo', 'Status', 'Região', 'Foco', 'Tipo Visita', 'Observação']
    registros_rows = [registros_header]

    for s, u in query_results:
        # data/hora safe formatting
        data_str = ''
        try:
            if getattr(s, 'data_agendamento', None):
                data_str = s.data_agendamento.strftime("%d/%m/%Y") if hasattr(s.data_agendamento, 'strftime') else str(s.data_agendamento)
            else:
                data_str = s.data_criacao.strftime("%d/%m/%Y") if hasattr(s.data_criacao, 'strftime') else str(s.data_criacao)
        except:
            data_str = str(getattr(s, 'data_agendamento', '') or getattr(s, 'data_criacao', ''))

        hora = getattr(s, 'hora_agendamento', '')
        hora_str = hora.strftime("%H:%M") if hasattr(hora, 'strftime') else str(hora or '')

        unidade = getattr(u, 'nome_uvis', '') or "Não informado"
        protocolo = getattr(s, 'protocolo', '') or ''
        status = getattr(s, 'status', '') or ''
        regiao = getattr(u, 'regiao', '') or ''
        foco = getattr(s, 'foco', '') or ''
        tipo_visita = getattr(s, 'tipo_visita', '') or ''
        obs = getattr(s, 'observacao', '') or ''

        registros_rows.append([data_str, hora_str, unidade, protocolo, status, regiao, foco, tipo_visita, obs])

    # Dividimos a tabela em pedaços para evitar problemas de memória/páginas
    # e garantir que não estoure
    chunk_size = 40
    for i in range(0, len(registros_rows), chunk_size):
        chunk = registros_rows[i:i+chunk_size]
        tbl = Table(chunk, repeatRows=1, colWidths=[18*mm, 14*mm, 35*mm, 26*mm, 22*mm, 28*mm, 28*mm, 30*mm, 45*mm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fbfdff')]),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 8))
        # adiciona quebra de página entre chunks (exceto se for o último)
        if i + chunk_size < len(registros_rows):
            story.append(PageBreak())

    # -------------------------
    # Footer fixo e page numbers
    # -------------------------
    # Usaremos canvas callbacks quando build() for chamado.
    def _header_footer(canvas, doc):
        # header (linha superior colorida)
        canvas.saveState()
        w, h = pagesize
        # linha azul
        canvas.setFillColor(colors.HexColor('#0d6efd'))
        canvas.rect(doc.leftMargin, h - (12*mm), doc.width, 4, fill=1, stroke=0)

        # rodapé: texto e número de página
        footer_text = "Sistema de Gestão de Solicitações — SGSV"
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor('#777'))
        canvas.drawString(doc.leftMargin, 10*mm, footer_text)

        # número de páginas
        page_num_text = f"Página {canvas.getPageNumber()}"
        canvas.drawRightString(doc.leftMargin + doc.width, 10*mm, page_num_text)
        canvas.restoreState()

    # -------------------------
    # Build e retorno
    # -------------------------
    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    # nome do arquivo
    nome_arquivo = f"relatorio_SGSV_{ano}_{mes:02d}"
    if uvis_id:
        nome_arquivo += f"_UVIS_{uvis_id}"

    # envia o pdf
    return send_file(
        caminho_pdf,
        as_attachment=True,
        download_name=f"{nome_arquivo}.pdf",
        mimetype="application/pdf"
    )

# =======================================================================
# ROTA 3: Exportar Excel (Com Filtro UVIS)
# =======================================================================
@bp.route('/admin/exportar_relatorio_excel')
def exportar_relatorio_excel():
    # 1. Parâmetros de Filtro
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    mes = request.args.get('mes', datetime.now().month, type=int)
    ano = request.args.get('ano', datetime.now().year, type=int)
    uvis_id = request.args.get('uvis_id', type=int) # NOVO FILTRO
    filtro_data = f"{ano}-{mes:02d}"

    # 2. Busca de Dados
    query_dados = db.session.query(
        Solicitacao.id,
        Solicitacao.status,
        Solicitacao.foco,
        Solicitacao.tipo_visita,
        Solicitacao.altura_voo,
        Solicitacao.data_agendamento,
        Solicitacao.hora_agendamento,
        Solicitacao.cep,
        Solicitacao.logradouro,
        Solicitacao.numero,
        Solicitacao.bairro,
        Solicitacao.cidade,
        Solicitacao.uf,
        Solicitacao.latitude,
        Solicitacao.longitude,
        Usuario.nome_uvis,
        Usuario.regiao
    ) \
        .join(Usuario, Usuario.id == Solicitacao.usuario_id) \
        .filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)

    # APLICAÇÃO DO NOVO FILTRO
    if uvis_id:
        query_dados = query_dados.filter(Solicitacao.usuario_id == uvis_id)

    dados = query_dados.all()

    # 3. Criar arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Cabeçalho
    colunas = [
        "ID", "Status", "Foco", "Tipo Visita", "Altura Voo",
        "Data Agendamento", "Hora Agendamento",
        "CEP", "Logradouro", "Número", "Bairro", "Cidade", "UF",
        "Latitude", "Longitude", "UVIS", "Região"
    ]

    # ... (Estilos e escrita do cabeçalho) ...
    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin', color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra1 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    zebra2 = PatternFill(start_color="FFF7FBFF", end_color="FFF7FBFF", fill_type="solid")

    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # 4. Preenchimento das linhas
    for row_num, row in enumerate(dados, 2):

        # ---- FORMATAR DATAS ----
        data_agendamento_fmt = ""
        if row.data_agendamento:
            try:
                data_agendamento_fmt = row.data_agendamento.strftime("%d/%m/%Y")
            except:
                data_agendamento_fmt = str(row.data_agendamento)

        # ---- FORMATAR HORA ----
        hora_agendamento_fmt = ""
        if row.hora_agendamento:
            try:
                hora_agendamento_fmt = row.hora_agendamento.strftime("%H:%M")
            except:
                hora_agendamento_fmt = str(row.hora_agendamento)

        # ---- PREENCHER LINHAS ----
        values = [
            row.id,
            row.status,
            row.foco,
            row.tipo_visita,
            row.altura_voo,
            data_agendamento_fmt,
            hora_agendamento_fmt,
            row.cep,
            row.logradouro,
            row.numero,
            row.bairro,
            row.cidade,
            row.uf,
            row.latitude,
            row.longitude,
            row.nome_uvis,
            row.regiao
        ]

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_index, value=value)
            cell.border = thin_border
            if col_index in (1, 3, 6, 8, 15, 16):
                cell.alignment = center
            else:
                cell.alignment = Alignment(vertical="top", horizontal="left")

            fill = zebra1 if (row_num % 2 == 0) else zebra2
            cell.fill = fill

    # 5. Ajustar e Finalizar
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max(10, min(max_length + 2, 60))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    # Gerar arquivo em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nome do arquivo
    nome_arquivo = f"relatorio_SGSV_{ano}_{mes:02d}"
    if uvis_id:
        nome_arquivo += f"_UVIS_{uvis_id}"

    return send_file(
        output,
        download_name=f"{nome_arquivo}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # NOVO: ROTA PARA EDIÇÃO COMPLETA (Apenas ADMIN)
@bp.route('/admin/editar_completo/<int:id>', methods=['GET', 'POST'])
def admin_editar_completo(id):
    # Permite APENAS 'admin'
    if session.get('user_tipo') != 'admin':
        flash('Acesso restrito. Apenas administradores podem editar detalhes do registro.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    pedido = Solicitacao.query.get_or_404(id)

    if request.method == 'POST':
        try:
            # 1. TRATAMENTO DE DATAS E HORAS
            data_str = request.form.get('data_agendamento')
            hora_str = request.form.get('hora_agendamento')

            data_obj = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            hora_obj = datetime.strptime(hora_str, '%H:%M').time() if hora_str else None

            # 2. TRATAMENTO DE BOOLEANOS
            criadouro_bool = request.form.get('criadouro') == 'sim'
            apoio_cet_bool = request.form.get('apoio_cet') == 'sim'

            # 3. ATUALIZAÇÃO DOS CAMPOS ORIGINAIS (Somente ADMIN)
            pedido.data_agendamento = data_obj
            pedido.hora_agendamento = hora_obj
            pedido.foco = request.form.get('foco')
            pedido.tipo_visita = request.form.get('tipo_visita')
            pedido.altura_voo = request.form.get('altura_voo')
            pedido.criadouro = criadouro_bool
            pedido.apoio_cet = apoio_cet_bool
            pedido.observacao = request.form.get('observacao')
            
            # Endereço
            pedido.cep = request.form.get('cep')
            pedido.logradouro = request.form.get('logradouro')
            pedido.bairro = request.form.get('bairro')
            pedido.cidade = request.form.get('cidade')
            pedido.uf = request.form.get('uf')
            pedido.numero = request.form.get('numero')
            pedido.complemento = request.form.get('complemento')

            # Protocolo / Status / GPS (também editáveis pelo ADMIN)
            pedido.protocolo = request.form.get('protocolo')
            pedido.status = request.form.get('status')
            pedido.justificativa = request.form.get('justificativa')
            pedido.latitude = request.form.get('latitude')
            pedido.longitude = request.form.get('longitude')
            
            db.session.commit()
            flash('Registro atualizado (ADMIN).', 'success')
            return redirect(url_for('main.admin_dashboard'))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar o registro: {e}", "danger")
            # Redireciona de volta para o GET com o ID para manter o contexto
            return redirect(url_for('main.admin_editar_completo', id=id)) 
    
    # GET: Exibe o formulário de edição (você precisará criar este template: 'admin_editar_completo.html')
    return render_template('admin_editar_completo.html', pedido=pedido)



from sqlalchemy.orm import joinedload
from flask import session, flash, redirect, url_for

@bp.route('/admin/deletar/<int:id>', methods=['POST'], endpoint='deletar_registro')
def deletar(id):

    if session.get('user_tipo') != 'admin':
        flash('Permissão negada. Apenas administradores podem deletar registros.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # Carrega o autor junto (evita lazy load pós-delete)
    pedido = (
        Solicitacao.query
        .options(joinedload(Solicitacao.autor))
        .get_or_404(id)
    )

    pedido_id = pedido.id
    autor_nome = pedido.autor.nome_uvis if pedido.autor else "UVIS"

    try:
        db.session.delete(pedido)
        db.session.commit()
    except Exception:
        db.session.rollback()
        # Não mostra erro ao usuário
        pass

    flash(f"Pedido #{pedido_id} da {autor_nome} deletado permanentemente.", "success")
    return redirect(url_for('main.admin_dashboard'))
# ----------------------------------------------
# ROTA DA AGENDA / CALENDÁRIO
# ----------------------------------------------
@bp.route("/agenda")
def agenda():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    user_tipo = session.get("user_tipo")
    user_id = session.get("user_id")

    # Admin, Operário e Visualizar enxergam tudo
    if user_tipo in ['admin', 'operario', 'visualizar']:
        eventos = Solicitacao.query.options(joinedload(Solicitacao.autor)).all()
    else:
        # UVIS vê apenas seus próprios agendamentos
        eventos = Solicitacao.query \
            .filter_by(usuario_id=user_id) \
            .options(joinedload(Solicitacao.autor)).all()

    # Converter eventos para o FullCalendar (JSON)
    agenda_eventos = []
    for e in eventos:
        if not e.data_agendamento:
            continue

        data = e.data_agendamento.strftime("%Y-%m-%d")
        hora = e.hora_agendamento.strftime("%H:%M") if e.hora_agendamento else "00:00"

        agenda_eventos.append({
            "title": f"{e.foco} - {e.autor.nome_uvis}",
            "start": f"{data}T{hora}",
            "url": url_for("main.admin_editar", id=e.id) if user_tipo in ["admin", "operario"] else None,
            "color": "#198754" if e.status == "APROVADO" else
                     "#dc3545" if e.status == "NEGADO" else
                     "#ffc107" if e.status == "EM ANÁLISE" else
                     "#0d6efd"
        })

    return render_template("agenda.html", eventos_json=json.dumps(agenda_eventos))